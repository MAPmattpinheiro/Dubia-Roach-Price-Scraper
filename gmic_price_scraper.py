"""
GMIC Competitor Price Scraper — v2.3
Scans the web for current dubia roach pricing across US and Canada.
Compares against GMIC's price list with per-roach estimates,
CAD/USD conversion, subscription tracking, stock status,
and price history appending.

Run: python gmic_price_scraper.py
Output: gmic_price_analysis.xlsx
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import time
import re
import os

# ─── GMIC BASELINE PRICES ────────────────────────────────────────────────────

GMIC_PRICES = {
    "live": {
        "Small / Nymphs (50-100ct)":  {"price": 12.99, "count_mid": 75,  "size": "1/4\"-3/8\""},
        "Medium (50-100ct)":          {"price": 17.99, "count_mid": 75,  "size": "1/2\"-5/8\""},
        "Large (25-50ct)":            {"price": 22.99, "count_mid": 37,  "size": "7/8\"-1\""},
        "Adult Males (50ct)":         {"price": 27.99, "count_mid": 50,  "size": "~1.5\""},
        "Adult Females (25ct)":       {"price": None,  "count_mid": 25,  "size": "~1.5\""},  # set price when GMIC offers this SKU
    },
    "freeze_dried": {
        "Standard Jar 1.3oz":         {"price": 13.99, "count_mid": 47,  "oz": 1.3},
        "Premium Pouch 2oz":          {"price": 18.99, "count_mid": 70,  "oz": 2.0},
        "Large Pouch 4oz":            {"price": 32.99, "count_mid": 140, "oz": 4.0},
        "Mixed Feeder Blend 2oz":     {"price": 19.99, "count_mid": None,"oz": 2.0},
    },
    "subscription": {
        "Starter (5% off monthly)":   {"price": None, "discount_pct": 5},
        "Keeper (10% off monthly)":   {"price": None, "discount_pct": 10},
        "Breeder (15% off biweekly)": {"price": None, "discount_pct": 15},
    },
    "free_ship_threshold": 100.00,
    "currency": "USD"
}

# ─── CAD/USD EXCHANGE RATE ───────────────────────────────────────────────────
# Fetched live at runtime via open.er-api.com (free, no API key needed)
# Falls back to hardcoded value if the API is unreachable

CAD_TO_USD_FALLBACK = 0.73  # fallback if API unavailable

def fetch_cad_usd_rate():
    """Fetch live CAD/USD rate from open.er-api.com (free, no key needed)."""
    try:
        resp = requests.get("https://open.er-api.com/v6/latest/CAD", timeout=8)
        data = resp.json()
        if data.get("result") == "success":
            rate = round(data["rates"]["USD"], 4)
            fetched_at = datetime.now().strftime("%Y-%m-%d %H:%M")
            print(f"   Live CAD/USD rate: {rate} (fetched {fetched_at})")
            return rate, fetched_at, "live"
        raise ValueError("API non-success")
    except Exception as e:
        print(f"   Live rate unavailable ({e}), using fallback: {CAD_TO_USD_FALLBACK}")
        return CAD_TO_USD_FALLBACK, datetime.now().strftime("%Y-%m-%d %H:%M"), "fallback"

# Set at runtime in main()
CAD_TO_USD      = CAD_TO_USD_FALLBACK
CAD_RATE_AT     = ""
CAD_RATE_SOURCE = "fallback"

# ─── COMPETITORS TO SCRAPE ───────────────────────────────────────────────────

COMPETITORS = [
    # ── US Live Feeder Sellers ──────────────────────────────────────────────
    {
        "name": "Dubia.com",
        "country": "US",
        "url": "https://dubiaroaches.com/products/dubia-roaches",
        "type": "live",
        "currency": "USD",
        "notes": "Largest volume seller, auctions model",
    },
    {
        "name": "Dubia Roach Depot",
        "country": "US",
        "url": "https://dubiaroachdepot.com/shop/buy-dubia-roaches",
        "type": "live",
        "currency": "USD",
        "notes": "Premium positioning, organic option available",
    },
    {
        "name": "The Critter Depot",
        "country": "US",
        "url": "https://www.thecritterdepot.com/products/dubia-roaches-critter-depot",
        "type": "live",
        "currency": "USD",
        "notes": "Multi-facility, fast shipping",
    },
    {
        "name": "Reptilian Arts",
        "country": "US",
        "url": "https://reptilianarts.com/products/dubia-roaches",
        "type": "live",
        "currency": "USD",
        "notes": "Small seller, Cumberland MD",
    },
    {
        "name": "TopFlight Dubia",
        "country": "US",
        "url": "https://topflightdubia.com/shop/",
        "type": "live",
        "currency": "USD",
        "notes": "Organic fed, premium market",
    },
    {
        "name": "Dubia Deli",
        "country": "US",
        "url": "https://www.dubideli.com/collections/dubia-roaches-for-sale",
        "type": "live",
        "currency": "USD",
        "notes": "Sells by count and by pound",
    },
    {
        "name": "ABDragons",
        "country": "US",
        "url": "https://abdragons.com/roaches/",
        "type": "live",
        "currency": "USD",
        "notes": "Bearded dragon focused",
    },
    {
        "name": "Dubia Roach Broker",
        "country": "US",
        "url": "https://dubiaroachbroker.com/collections/all",
        "type": "live",
        "currency": "USD",
        "notes": "Sells by gram/pound — bulk model",
    },
    {
        "name": "Backwater Reptiles",
        "country": "US",
        "url": "https://www.backwaterreptiles.com/feeders/dubia-roaches-for-sale.html",
        "type": "live",
        "currency": "USD",
        "notes": "Large reptile retailer",
    },
    {
        "name": "Chewy - Live Feeders",
        "country": "US",
        "url": "https://news.google.com/rss/search?q=chewy+dubia+roaches+price&hl=en-US&gl=US&ceid=US:en",
        "type": "live",
        "currency": "USD",
        "notes": "Major retail channel — via Google News RSS (site blocks scrapers)",
        "use_rss": True,
    },
    # ── US Freeze-Dried Sellers ─────────────────────────────────────────────
    {
        "name": "Fluker's Direct",
        "country": "US",
        "url": "https://flukerfarms.com/Freeze-Dried-Dubia-Roaches",
        "type": "freeze_dried",
        "currency": "USD",
        "notes": "Primary benchmark competitor",
    },
    {
        "name": "Chewy - Fluker's FD",
        "country": "US",
        "url": "https://news.google.com/rss/search?q=chewy+flukers+freeze+dried+dubia+roaches&hl=en-US&gl=US&ceid=US:en",
        "type": "freeze_dried",
        "currency": "USD",
        "notes": "Retail channel pricing — via Google News RSS (site blocks scrapers)",
        "use_rss": True,
    },
    {
        "name": "Walmart - Fluker's FD",
        "country": "US",
        "url": "https://www.walmart.com/ip/Flukers-Freeze-Dried-Dubia-Roaches-for-Reptiles/3119991813",
        "type": "freeze_dried",
        "currency": "USD",
        "notes": "Mass retail channel pricing",
    },
    {
        "name": "Petco - FD Feeders",
        "country": "US",
        "url": "https://news.google.com/rss/search?q=petco+freeze+dried+dubia+roaches+price&hl=en-US&gl=US&ceid=US:en",
        "type": "freeze_dried",
        "currency": "USD",
        "notes": "Major pet retail chain — via Google News RSS (site blocks scrapers)",
        "use_rss": True,
    },
    {
        "name": "PetSmart - FD Feeders",
        "country": "US",
        "url": "https://www.petsmart.com/reptile/food-and-supplements/",
        "type": "freeze_dried",
        "currency": "USD",
        "notes": "Major pet retail chain",
    },
    {
        "name": "Amazon US - FD Dubia",
        "country": "US",
        "url": "https://news.google.com/rss/search?q=amazon+freeze+dried+dubia+roaches+price&hl=en-US&gl=US&ceid=US:en",
        "type": "freeze_dried",
        "currency": "USD",
        "notes": "Marketplace pricing — via Google News RSS (site blocks scrapers)",
        "use_rss": True,
    },
    # ── Subscription Tracking ───────────────────────────────────────────────
    {
        "name": "Dubia.com Subscription",
        "country": "US",
        "url": "https://dubiaroaches.com/collections/subscriptions",
        "type": "subscription",
        "currency": "USD",
        "notes": "Recurring delivery pricing",
    },
    {
        "name": "Dubia Roach Depot Sub",
        "country": "US",
        "url": "https://dubiaroachdepot.com/shop/buy-dubia-roaches",
        "type": "subscription",
        "currency": "USD",
        "notes": "Recurring shipment pricing",
    },
    # ── Canada Sellers ──────────────────────────────────────────────────────
    {
        "name": "Amazon.ca - FD Dubia",
        "country": "CA",
        "url": "https://www.amazon.ca/s?k=freeze+dried+dubia+roaches",
        "type": "freeze_dried",
        "currency": "CAD",
        "notes": "Canadian marketplace — key competitor for export strategy",
    },
    {
        "name": "PetSmart Canada - FD",
        "country": "CA",
        "url": "https://www.petsmart.ca/reptile/food-and-supplements/",
        "type": "freeze_dried",
        "currency": "CAD",
        "notes": "Canadian retail chain",
    },
    {
        "name": "Pisces Pros Canada",
        "country": "CA",
        "url": "https://www.piscespros.com/collections/feeders",
        "type": "live",
        "currency": "CAD",
        "notes": "Major Canadian reptile supplier",
    },
    {
        "name": "Big Al's Canada",
        "country": "CA",
        "url": "https://www.bigalspets.ca/reptiles/reptile-food/",
        "type": "freeze_dried",
        "currency": "CAD",
        "notes": "Canadian pet chain",
    },
    {
        "name": "Josh's Frogs Canada",
        "country": "CA",
        "url": "https://www.joshsfrogs.com/dubia-roaches.html",
        "type": "live",
        "currency": "CAD",
        "notes": "US seller shipping to Canada",
    },
]

# ─── PRICE PATTERNS ───────────────────────────────────────────────────────────

PRICE_PATTERNS = [
    r'\$\s*(\d+\.?\d*)',
    r'USD\s*(\d+\.?\d*)',
    r'CAD\s*(\d+\.?\d*)',
    r'(\d+\.?\d*)\s*(?:USD|CAD)',
]

SHIP_KEYWORDS = ["free shipping", "free ship", "ships free", "free delivery"]

OUT_OF_STOCK_KEYWORDS = [
    "out of stock", "sold out", "unavailable", "notify me",
    "currently unavailable", "back order", "backordered"
]

SUBSCRIPTION_KEYWORDS = [
    "subscribe", "subscription", "recurring", "auto-ship",
    "autoship", "subscribe & save", "subscribe and save"
]

SIZE_LABELS = {
    "small":           "Small",
    "nymph":           "Small",
    "1/4":             "Small",
    "3/8":             "Small",
    "medium":          "Medium",
    "1/2":             "Medium",
    "5/8":             "Medium",
    "large":           "Large",
    "7/8":             "Large",
    "adult female":    "Adult Female",
    "female dubia":    "Adult Female",
    "female roach":    "Adult Female",
    "adult male":      "Adult Male",
    "male dubia":      "Adult Male",
    "male roach":      "Adult Male",
    "adult":           "Adult",
    "xl":              "XL / Adult",
    "jumbo":           "XL / Adult",
    "pre-adult":       "XL / Adult",
    "1.5":             "Adult",
    "1.3":             "1.3oz Jar",
    "1.3 oz":          "1.3oz Jar",
    "2 oz":            "2oz Pouch",
    "4 oz":            "4oz Pouch",
    "2oz":             "2oz Pouch",
    "4oz":             "4oz Pouch",
    "16 oz":           "16oz Bulk",
    "16oz":            "16oz Bulk",
    "bulk":            "Bulk",
    "per pound":       "By Pound",
    "per gram":        "By Gram",
}

# Estimated roach counts per size for per-roach calculation
COUNT_ESTIMATES = {
    "Small":        75,
    "Medium":       75,
    "Large":        37,
    "Adult":        50,
    "Adult Male":   50,
    "Adult Female": 25,
    "XL / Adult":   25,
    "1.3oz Jar":    47,
    "2oz Pouch":    70,
    "4oz Pouch":    140,
    "16oz Bulk":    500,
}

# ─── SCRAPER ─────────────────────────────────────────────────────────────────

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

def fetch_page(url):
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        return BeautifulSoup(resp.text, "lxml")
    except Exception as e:
        print(f"     Could not fetch {url}: {e}")
        return None

def extract_prices_from_text(text):
    prices = []
    for pattern in PRICE_PATTERNS:
        for m in re.findall(pattern, text):
            try:
                val = float(m)
                if 1.0 <= val <= 500.0:
                    prices.append(val)
            except:
                pass
    return list(set(prices))

def find_free_shipping(soup):
    text = soup.get_text().lower()
    for kw in SHIP_KEYWORDS:
        if kw in text:
            match = re.search(
                r'free\s+ship(?:ping)?\s+(?:on\s+orders?\s+)?(?:over|above|of)?\s*\$?\s*(\d+)',
                text, re.I
            )
            if match:
                return float(match.group(1))
            return 0.0
    return None

def detect_out_of_stock(soup):
    text = soup.get_text().lower()
    return any(kw in text for kw in OUT_OF_STOCK_KEYWORDS)

def detect_subscription(soup):
    text = soup.get_text().lower()
    if any(kw in text for kw in SUBSCRIPTION_KEYWORDS):
        match = re.search(
            r'(\d+)\s*%\s*off\s*(?:with\s*)?(?:subscription|subscribe)',
            text, re.I
        )
        if match:
            return f"{match.group(1)}% off"
        return "Yes (% unknown)"
    return "No"

def categorize_price_context(context_text):
    ctx = context_text.lower()
    for key, label in SIZE_LABELS.items():
        if key in ctx:
            return label
    return "Unknown"

def per_roach_price(price, size_label, currency):
    count = COUNT_ESTIMATES.get(size_label)
    if count and price:
        usd = price * CAD_TO_USD if currency == "CAD" else price
        return round(usd / count, 3)
    return None

def scrape_rss_fallback(comp):
    """For sites that block scrapers, fetch Google News RSS and extract any price mentions."""
    print(f"  Scraping {comp['name']} ({comp['country']}) via RSS fallback...")
    try:
        resp = requests.get(comp["url"], headers=HEADERS, timeout=12)
        soup = BeautifulSoup(resp.content, "xml")
        items = soup.find_all("item")
        rows = []
        seen = set()
        for item in items:
            title = item.find("title")
            desc  = item.find("description")
            text  = (title.get_text() if title else "") + " " + (desc.get_text() if desc else "")
            prices = extract_prices_from_text(text)
            for p in prices:
                key = round(p, 2)
                if key in seen:
                    continue
                seen.add(key)
                category = categorize_price_context(text)
                price_usd = round(p * CAD_TO_USD, 2) if comp["currency"] == "CAD" else p
                per_roach = per_roach_price(p, category, comp["currency"])
                rows.append({
                    "Scraped Date":      datetime.now().strftime("%Y-%m-%d"),
                    "Competitor":        comp["name"],
                    "Country":           comp["country"],
                    "Product Type":      {"live": "Live Feeder", "freeze_dried": "Freeze-Dried",
                                          "subscription": "Subscription"}.get(comp["type"], comp["type"]),
                    "Size / Product":    category,
                    "Price (Native)":    p,
                    "Currency":          comp["currency"],
                    "Price (USD Est.)":  price_usd,
                    "Per-Roach (USD)":   per_roach,
                    "Free Ship At":      "Unknown (RSS)",
                    "Subscription":      "N/A",
                    "In Stock":          "Unknown",
                    "CAD/USD Rate":      CAD_TO_USD,
                    "Rate Source":       CAD_RATE_SOURCE,
                    "Rate Fetched At":   CAD_RATE_AT,
                    "Notes":             comp.get("notes", "") + " | Price from news mention",
                    "Source URL":        comp["url"],
                    "Context Snippet":   text[:120].strip(),
                })
        print(f"     {len(rows)} price mentions found via RSS")
        return rows
    except Exception as e:
        print(f"     RSS fallback failed: {e}")
        return []

def scrape_competitor(comp):
    # Use RSS fallback for sites known to block scrapers
    if comp.get("use_rss"):
        return scrape_rss_fallback(comp)

    print(f"  Scraping {comp['name']} ({comp['country']})...")
    soup = fetch_page(comp["url"])
    if not soup:
        return []

    page_text  = soup.get_text(separator=" ")
    prices     = extract_prices_from_text(page_text)
    free_ship  = find_free_shipping(soup)
    oos        = detect_out_of_stock(soup)
    sub_info   = detect_subscription(soup) if comp["type"] == "subscription" else "N/A"

    if not prices:
        print(f"     No prices found")
        return []

    # Structured price extraction
    price_elements = soup.find_all(
        class_=re.compile(r'price|cost|amount|money', re.I)
    )

    found_pairs = []
    for el in price_elements:
        el_text = el.get_text(strip=True)
        for pm in re.findall(r'\$\s*(\d+\.?\d*)', el_text):
            try:
                val = float(pm)
                if 1.0 <= val <= 500.0:
                    parent_text = el.parent.get_text(separator=" ", strip=True) if el.parent else el_text
                    category = categorize_price_context(parent_text or el_text)
                    found_pairs.append((val, category, parent_text[:150]))
            except:
                pass

    if not found_pairs:
        for p in prices[:10]:
            found_pairs.append((p, "See URL", ""))

    # Deduplicate
    seen = set()
    rows = []
    for price, category, context in found_pairs:
        key = (round(price, 2), category)
        if key in seen:
            continue
        seen.add(key)

        price_usd = round(price * CAD_TO_USD, 2) if comp["currency"] == "CAD" else price
        per_roach = per_roach_price(price, category, comp["currency"])

        rows.append({
            "Scraped Date":      datetime.now().strftime("%Y-%m-%d"),
            "Competitor":        comp["name"],
            "Country":           comp["country"],
            "Product Type":      {"live": "Live Feeder", "freeze_dried": "Freeze-Dried",
                                  "subscription": "Subscription"}.get(comp["type"], comp["type"]),
            "Size / Product":    category,
            "Price (Native)":    price,
            "Currency":          comp["currency"],
            "Price (USD Est.)":  price_usd,
            "Per-Roach (USD)":   per_roach,
            "Free Ship At":      (f"${free_ship:.0f}" if free_ship and free_ship > 0
                                  else ("Yes (no min)" if free_ship == 0 else "No / Unknown")),
            "Subscription":      sub_info,
            "In Stock":          "No" if oos else "Yes",
            "CAD/USD Rate":      CAD_TO_USD,
            "Rate Source":       CAD_RATE_SOURCE,
            "Rate Fetched At":   CAD_RATE_AT,
            "Notes":             comp.get("notes", ""),
            "Source URL":        comp["url"],
            "Context Snippet":   context.strip(),
        })

    print(f"     {len(rows)} price points | Stock: {'OUT' if oos else 'OK'} | Sub: {sub_info}")
    return rows

# ─── COMPARISON ENGINE ────────────────────────────────────────────────────────

def compare_to_gmic(df):
    def get_gmic_price(row):
        ptype = row["Product Type"]
        size  = row["Size / Product"].lower()
        if ptype == "Live Feeder":
            if "small" in size or "nymph" in size or "1/4" in size:
                return GMIC_PRICES["live"]["Small / Nymphs (50-100ct)"]["price"]
            elif "medium" in size or "1/2" in size:
                return GMIC_PRICES["live"]["Medium (50-100ct)"]["price"]
            elif "large" in size or "7/8" in size:
                return GMIC_PRICES["live"]["Large (25-50ct)"]["price"]
            elif "adult female" in size or "female" in size:
                return GMIC_PRICES["live"]["Adult Females (25ct)"]["price"]  # None until GMIC offers this
            elif "adult male" in size or "adult" in size or "xl" in size or "1.5" in size:
                return GMIC_PRICES["live"]["Adult Males (50ct)"]["price"]
        elif ptype == "Freeze-Dried":
            if "1.3" in size or "jar" in size:
                return GMIC_PRICES["freeze_dried"]["Standard Jar 1.3oz"]["price"]
            elif "2oz" in size or "2 oz" in size:
                return GMIC_PRICES["freeze_dried"]["Premium Pouch 2oz"]["price"]
            elif "4oz" in size or "4 oz" in size:
                return GMIC_PRICES["freeze_dried"]["Large Pouch 4oz"]["price"]
        return None

    def get_gmic_per_roach(row):
        ptype = row["Product Type"]
        size  = row["Size / Product"].lower()
        if ptype == "Live Feeder":
            if "small" in size or "nymph" in size:
                p = GMIC_PRICES["live"]["Small / Nymphs (50-100ct)"]
                return round(p["price"] / p["count_mid"], 3) if p["price"] else None
            elif "medium" in size:
                p = GMIC_PRICES["live"]["Medium (50-100ct)"]
                return round(p["price"] / p["count_mid"], 3)
            elif "large" in size:
                p = GMIC_PRICES["live"]["Large (25-50ct)"]
                return round(p["price"] / p["count_mid"], 3)
            elif "adult female" in size or "female" in size:
                return None  # GMIC doesn't offer this SKU yet — shows gap vs competitors
            elif "adult male" in size or "adult" in size or "xl" in size:
                p = GMIC_PRICES["live"]["Adult Males (50ct)"]
                return round(p["price"] / p["count_mid"], 3)
        elif ptype == "Freeze-Dried":
            if "1.3" in size or "jar" in size:
                p = GMIC_PRICES["freeze_dried"]["Standard Jar 1.3oz"]
                return round(p["price"] / p["count_mid"], 3)
            elif "2oz" in size or "2 oz" in size:
                p = GMIC_PRICES["freeze_dried"]["Premium Pouch 2oz"]
                return round(p["price"] / p["count_mid"], 3)
            elif "4oz" in size or "4 oz" in size:
                p = GMIC_PRICES["freeze_dried"]["Large Pouch 4oz"]
                return round(p["price"] / p["count_mid"], 3)
        return None

    def assess(row):
        gmic = row["GMIC Price (USD)"]
        comp = row["Price (USD Est.)"]
        if gmic is None or comp is None:
            return "N/A"
        diff_pct = ((gmic - comp) / comp) * 100
        if diff_pct > 15:
            return f"HIGH: GMIC {diff_pct:.0f}% higher"
        elif diff_pct < -15:
            return f"LOW: GMIC {abs(diff_pct):.0f}% lower"
        return "OK: Competitive"

    df["GMIC Price (USD)"]      = df.apply(get_gmic_price, axis=1)
    df["GMIC Per-Roach (USD)"]  = df.apply(get_gmic_per_roach, axis=1)
    df["vs GMIC"]               = df.apply(assess, axis=1)
    return df

# ─── EXCEL OUTPUT ─────────────────────────────────────────────────────────────

def style_header(ws, col_count):
    hfill  = PatternFill("solid", start_color="1F4E79")
    hfont  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    halign = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),  bottom=Side(style="thin", color="D9D9D9"),
    )
    ws.row_dimensions[1].height = 22
    for i in range(1, col_count + 1):
        c = ws.cell(row=1, column=i)
        c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = border

def style_rows(ws, col_count):
    body  = Font(name="Arial", size=10)
    ctr   = Alignment(horizontal="center", vertical="center")
    left  = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),  bottom=Side(style="thin", color="D9D9D9"),
    )
    vs_colors = {"OK:": "C6EFCE", "HIGH:": "FFC7CE", "LOW:": "FFEB9C"}
    oos_cols = {}

    # Find column indices by header
    headers = [ws.cell(row=1, column=i).value for i in range(1, col_count + 1)]
    vs_col  = headers.index("vs GMIC") + 1 if "vs GMIC" in headers else None
    oos_col = headers.index("In Stock") + 1 if "In Stock" in headers else None

    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 22
        for col_idx, cell in enumerate(row, 1):
            cell.font = body; cell.border = border
            cell.alignment = ctr if col_idx in [1,2,3,4,6,7,8,9,10,11,12] else left
            if vs_col and col_idx == vs_col:
                val = str(cell.value or "")
                for emoji, color in vs_colors.items():
                    if emoji in val:
                        cell.fill = PatternFill("solid", start_color=color)
                        break
            if oos_col and col_idx == oos_col:
                if str(cell.value) == "No":
                    cell.fill = PatternFill("solid", start_color="FFC7CE")

def append_history(df, filepath):
    """Append today's data to a history sheet for trend tracking."""
    if not os.path.exists(filepath):
        return
    try:
        wb = load_workbook(filepath)
        if "Price History" not in wb.sheetnames:
            ws = wb.create_sheet("Price History")
            ws.append(list(df.columns))
        else:
            ws = wb["Price History"]
        for row in df.itertuples(index=False):
            ws.append(list(row))
        wb.save(filepath)
        print(f"  History sheet updated ({len(df)} rows appended)")
    except Exception as e:
        print(f"   Could not update history: {e}")

def write_excel(df, filepath="gmic_price_analysis.xlsx"):
    # Save history before overwriting main sheets
    append_history(df, filepath)

    wb = Workbook()

    # ── Competitor Prices ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Competitor Prices"
    ws1.freeze_panes = "A2"
    col_widths = [13,22,10,16,18,14,10,14,14,14,14,10,20,45,50,14,14,14]
    for i, w in enumerate(col_widths[:len(df.columns)], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws1.append(list(row))
    style_header(ws1, len(df.columns))
    style_rows(ws1, len(df.columns))

    # ── GMIC Baseline ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("GMIC Baseline")
    ws2.append(["Product Type", "Size / Product", "Price (USD)", "Est. Count", "Per-Roach (USD)"])
    for size, d in GMIC_PRICES["live"].items():
        per = round(d["price"] / d["count_mid"], 3) if d["count_mid"] else ""
        ws2.append(["Live Feeder", size, d["price"], d["count_mid"], per])
    for prod, d in GMIC_PRICES["freeze_dried"].items():
        per = round(d["price"] / d["count_mid"], 3) if d["count_mid"] else ""
        ws2.append(["Freeze-Dried", prod, d["price"], d["count_mid"], per])
    for plan, d in GMIC_PRICES["subscription"].items():
        ws2.append(["Subscription", plan, "See retail", "", f"{d['discount_pct']}% off"])
    ws2.append(["Shipping", f"Free at ${GMIC_PRICES['free_ship_threshold']:.0f}", "", "", ""])
    ws2.append(["CAD/USD Rate", CAD_TO_USD, CAD_RATE_SOURCE, CAD_RATE_AT, ""])
    for col, w in zip(["A","B","C","D","E"], [18,30,14,12,16]):
        ws2.column_dimensions[col].width = w
    style_header(ws2, 5)

    # ── Summary ──────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.append(["GMIC Competitor Price Analysis — Summary"])
    ws3.append([f"Run date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws3.append([f"CAD/USD rate: {CAD_TO_USD} ({CAD_RATE_SOURCE}, fetched {CAD_RATE_AT})"])
    ws3.append([])
    ws3.append(["Metric", "Value"])
    ws3.append(["Competitors scraped",      df["Competitor"].nunique()])
    ws3.append(["US competitors",           df[df["Country"]=="US"]["Competitor"].nunique()])
    ws3.append(["Canada competitors",       df[df["Country"]=="CA"]["Competitor"].nunique()])
    ws3.append(["Total price points",       len(df)])
    ws3.append(["Live feeder prices",       len(df[df["Product Type"]=="Live Feeder"])])
    ws3.append(["Freeze-dried prices",      len(df[df["Product Type"]=="Freeze-Dried"])])
    ws3.append(["Subscription data points", len(df[df["Product Type"]=="Subscription"])])
    ws3.append(["Out of stock detected",    (df["In Stock"]=="No").sum()])
    ws3.append([])
    ws3.append(["GMIC Competitive Status", "Count"])
    for label, emoji in [(" Competitive",""),(" GMIC Higher",""),(" GMIC Lower","")]:
        count = df["vs GMIC"].str.contains(emoji, na=False).sum() if "vs GMIC" in df.columns else 0
        ws3.append([label, count])
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 16
    style_header(ws3, 2)

    # ── Price History placeholder (created by append_history on next run) ────
    ws4 = wb.create_sheet("Price History")
    ws4.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws4.append(list(row))
    style_header(ws4, len(df.columns))

    wb.save(filepath)
    print(f"\n Saved: {filepath}")
    print(f"   {len(df)} price points | {df['Competitor'].nunique()} competitors")

# ─── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 60)
    print("  GMIC Competitor Price Scraper v2.3")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"  {len(COMPETITORS)} competitors")
    print("=" * 60 + "\n")

    # Fetch live CAD/USD rate first
    print("Fetching live CAD/USD exchange rate...")
    CAD_TO_USD, CAD_RATE_AT, CAD_RATE_SOURCE = fetch_cad_usd_rate()
    print()

    all_rows = []
    for comp in COMPETITORS:
        rows = scrape_competitor(comp)
        all_rows.extend(rows)
        time.sleep(1.5)

    if not all_rows:
        print("\n No data scraped. Check your network connection.")
    else:
        df = pd.DataFrame(all_rows)
        df = compare_to_gmic(df)
        df = df.sort_values(["Product Type", "Country", "Competitor"])
        write_excel(df)

        print("\nDone! Open gmic_price_analysis.xlsx")
        print(f"  CAD/USD rate used: {CAD_TO_USD} ({CAD_RATE_SOURCE})")
        competitive = df['vs GMIC'].str.startswith('OK', na=False).sum()
        higher      = df['vs GMIC'].str.startswith('HIGH', na=False).sum()
        lower       = df['vs GMIC'].str.startswith('LOW', na=False).sum()
        na          = df['vs GMIC'].eq('N/A').sum()
        print(f"\n  Competitive:    {competitive}")
        print(f"  GMIC Higher:    {higher}")
        print(f"  GMIC Lower:     {lower}")
        print(f"  Not Comparable: {na}")
        oos = (df["In Stock"] == "No").sum()
        if oos:
            print(f"\n  WARNING: {oos} out-of-stock detected — potential demand opportunity!")
